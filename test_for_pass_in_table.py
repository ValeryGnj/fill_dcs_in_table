from docxtpl import DocxTemplate
import openpyxl


def main():

    wb = openpyxl.load_workbook('data_staff_pass_fishport.xlsx', data_only=True)
    sheet = wb['список_людей']

    data_list = []
    for i in range(2, sheet.max_row + 1):
        data_list.append({'id': i-1,
                          'Name': sheet.cell(i, 2).value,
                          'Date_of_birth': sheet.cell(i, 3).value,
                          'place_of_birth': sheet.cell(i, 4).value,
                          'citezenship': sheet.cell(i, 5).value,
                          'passport_series': sheet.cell(i, 6).value,
                          'passport_numbers': sheet.cell(i, 7).value,
                          'issuing_authority_and_date': sheet.cell(i, 8).value,
                          'registration': sheet.cell(i, 9).value,
                          'rank': sheet.cell(i, 10).value})

    context = {'data_list': data_list}
    doc = DocxTemplate('template_order_staff_pass_fishport.docx')
    doc.render(context)
    doc.save('data_list_result.docx')


if __name__ == '__main__':
    main()
